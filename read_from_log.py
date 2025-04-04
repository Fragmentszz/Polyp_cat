# [2025-02-03 10:58:57 PM-eval_group.py-INFO:Testing on CVC-300 dataset...]
# [2025-02-03 10:59:23 PM-eval_group.py-INFO:Mean val dice: 0.9418825507164001]
# [2025-02-03 10:59:23 PM-eval_group.py-INFO:Mean val gd: 0.9418825437625249]
# [2025-02-03 10:59:23 PM-eval_group.py-INFO:Mean val iou: 0.8924141625563303]
# [2025-02-03 10:59:23 PM-eval_group.py-INFO:Testing on CVC-ClinicDB dataset...]
# [2025-02-03 10:59:48 PM-eval_group.py-INFO:Mean val dice: 0.94096619563718]
# [2025-02-03 10:59:48 PM-eval_group.py-INFO:Mean val gd: 0.94096619563718]
# [2025-02-03 10:59:48 PM-eval_group.py-INFO:Mean val iou: 0.8903281313757743]
# [2025-02-03 10:59:48 PM-eval_group.py-INFO:Testing on CVC-ColonDB dataset...]
# [2025-02-03 11:02:39 PM-eval_group.py-INFO:Mean val dice: 0.9239757787240179]
# [2025-02-03 11:02:39 PM-eval_group.py-INFO:Mean val gd: 0.9239757779397463]
# [2025-02-03 11:02:39 PM-eval_group.py-INFO:Mean val iou: 0.8639824384137204]
# [2025-02-03 11:02:39 PM-eval_group.py-INFO:Testing on ETIS-LaribPolypDB dataset...]
# [2025-02-03 11:04:45 PM-eval_group.py-INFO:Mean val dice: 0.9397643135518444]
# [2025-02-03 11:04:45 PM-eval_group.py-INFO:Mean val gd: 0.9397643153764763]
# [2025-02-03 11:04:45 PM-eval_group.py-INFO:Mean val iou: 0.8891609347596461]
# [2025-02-03 11:04:45 PM-eval_group.py-INFO:Testing on Kvasir dataset...]
# [2025-02-03 11:05:35 PM-eval_group.py-INFO:Mean val dice: 0.9593014198541642]
# [2025-02-03 11:05:35 PM-eval_group.py-INFO:Mean val gd: 0.9593014216423035]
# [2025-02-03 11:05:35 PM-eval_group.py-INFO:Mean val iou: 0.9233735775947571]

import re
from unittest import result


def read_from_text(lines):
    dataset_dict = {
        'Kvasir':'Kvasir',
        'CVC-ClinicDB':'CVC-ClinicDB',
        'CVC-ColonDB':'CVC-ColonDB',
        'CVC-300':'CVC-300',
        'ETIS-LaribPolypDB':'ETIS'
    }

    dataset = None
    iou = None
    gd = None
    dice = None
    results = {}
    for line in lines:
        if 'Testing on' in line:
            dataset = line.split(' ')[-2]
        if 'Mean val dice' in line:
            dice = float(line.split(' ')[-1].strip(']\n'))
        if 'Mean val gd' in line:
            gd = float(line.split(' ')[-1].strip(']\n'))
        if 'Mean val iou' in line:
            iou = float(line.split(' ')[-1].strip(']\n'))
        if dataset is not None and iou is not None and gd is not None and dice is not None:
            results[dataset_dict[dataset]] = (iou,gd,dice)
            dataset = None
            iou = None
            gd = None
            dice = None
    return results

def get_last_CVC_300_lines(lines):

    for i in range(len(lines)-1,0,-1):
        if 'Start Eval' in lines[i]:
            return lines[i+1:]
        elif 'Testing on CVC-300 dataset' in lines[i]:
            return lines[i:]
    raise ValueError('No CVC-300 dataset found in the log file')
    
def read_from_log(path):

    with open(path,'r') as f:
        lines = f.readlines()
        
        lines = get_last_CVC_300_lines(lines)
        return read_from_text(lines)
import os
import openpyxl

def write_to_excel(results,src_path,save_path):
    datasets = ['Kvasir','CVC-ClinicDB','CVC-ColonDB','CVC-300','ETIS']
    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    row = ws.max_row + 1
    for model in results:
        ws.cell(row=row,column=1,value=model)
        for i in range(5):
            ws[f'{chr(65+i*2+1)}{row}'] = results[model][datasets[i]][2]
            ws[f'{chr(65+i*2+2)}{row}'] = results[model][datasets[i]][0]
        row += 1
    wb.save(save_path)



if __name__ == '__main__':
    group_eval_model = ['6_1','6_3','6_7','7_3','7-5','7-6','7-7','8_2']
    eval_root = r'/remote-home/results'
    weights = ['connect_hq_token','local_layer','evp']
    results = {}
    for model in group_eval_model:
        p = int(model.split('_')[1])
        model_name = ''
        for i in range(3):
            if p % 2 == 1:
                model_name += weights[i] + '+'
            p = p // 2
        model_name = model + '_' + model_name.strip('+')
        for weight in weights:
            res = read_from_log(os.path.join(eval_root,model,'eval.log'))
            results[model_name] = res
            print(model_name)
            for key in res:
                print(key,res[key])
    write_to_excel(results,os.path.join(eval_root,'result.xlsx'),os.path.join(eval_root,'result2.xlsx'))
    
